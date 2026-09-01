#!/usr/bin/env python3
"""
This script has been created to help interpret the top features provided as the output of MicroICS from the “Train models” option. It reads the feature_ranking_label and feature_ranking_date files generated for the random forest-based training.

It produces an Excel workbook (plus a plain TSV) that gathers, for the top N (usually 100)
label features:
  - rank and importance score (from rankings_label)
  - base feature name and aggregation type (parsed from the feature string)
  - feature category, per-layer/image, and description (from Table S4 in Janež, Škrlj, Osojnik et al. 2026 (doi: 10.1038/s41522-026-01083-8))
  - per-class median and IQR for the feature value (from datafile.tsv, split by the “label” column)
  - whether the feature also ranks highly for date (batch-effect flag)
  - correlation-cluster membership (features with |r| > threshold cluster together, usually 0.8)
  - a plain-language hint to scaffold hypothesis-building (not a conclusion)

In the raw output, the rows are sorted by cluster, and each row is tinted with its cluster’s colour. The number in a cell is the value; the colour of the row indicates the cluster. To examine redundancy, sort the results within each cluster by aggregation type to determine whether the cluster can be interpreted as a whole pointing to one feature or to several.

To run this script you need Python installed https://www.python.org/downloads/ and you need to run the script using Terminal and the command line. You open the Terminal either by right-clicking → Open in Terminal or by searching for PowerShell. When the Terminal opens, you type the command provided below. It is also advisable to put all required files in one folder and run the script in that directory.

Usage:
python interpret_features.py --datafile datafile.tsv --label rankings_label.tsv --date rankings_date.tsv --n 100  --date-n 100 --corr-threshold 0.8 --out-prefix microics_feature_interpretation

Explanations:
--datafile: file with values of calculated features for the training dataset 
--label: feature ranking file for the label 
--date: feature ranking file for the date 
--n: number of top features to be analysed in the file in --label 
--date-n: number of top features to be analysed in the file in --date 
--corr-threshold: threshold for the correlation analysis 
--out-prefix: prefix for the output file names

"""

import argparse
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd


# ----------------------------------------------------------------------
# TABLE S4 LOOKUP  (base datafile stem -> category / per / description)
# ----------------------------------------------------------------------
S4 = {
    # ---- Morphological: counts ----
    "counts(inten<":        ("Morphological (counts)", "layer",
        "Number of bacteria per layer after thresholding and segmentation. Reflects abundance-related structure; conflates object counts with signal strength, so not an absolute abundance measure."),
    "counts(Norminten<":    ("Morphological (counts)", "layer",
        "Number of bacteria per layer using intensity-normalised input; counts expressed relative to the per-image total."),
    # ---- Morphological: volume / coverage ----
    "BioVolumeThr":         ("Morphological (volume)", "image",
        "Biofilm volume: thresholded biofilm pixels multiplied by voxel size. Total volume occupied by the biofilm."),
    "GPTVolumeNormalized":  ("Morphological (volume)", "layer",
        "Layer-wise biovolume, intensity-normalised, then aggregated per image."),
    "GPTVolume":            ("Morphological (volume)", "layer",
        "Layer-wise biovolume (thresholded per layer, then aggregated per image)."),
    "SubstratumRelativeCoverage": ("Morphological (coverage)", "image",
        "Percentage of the well surface (first layer) covered by biofilm."),
    # ---- Intensity: global / diff ----
    "globalMean":           ("Intensity (global)", "image",
        "Average pixel intensity of the whole 3D image; overall brightness/density."),
    "maxdiffs":             ("Intensity (layer-wise global)", "image",
        "Largest change in mean brightness between consecutive layers."),
    "mindiffs":             ("Intensity (layer-wise global)", "image",
        "Smallest change in mean brightness between layers; stable/low-change regions."),
    "mdiffs":               ("Intensity (layer-wise global)", "image",
        "Mean brightness difference between successive layers; vertical uniformity/variability."),
    "eigen":                ("Intensity (global)", "/",
        "A value representing overall 3D differences (noted but not currently calculated)."),
    "diffNormalized":       ("Intensity (contrast)", "layer",
        "Normalised max-minus-min intensity per layer; relative contrast distribution."),
    "diff":                 ("Intensity (contrast)", "layer",
        "Max-minus-min pixel intensity within each layer; contrast (mean), typical value (median), contrast variability (std), skew (percentiles)."),
    # ---- Intensity: summary statistics per layer ----
    "meanNormalized":       ("Intensity (mean)", "layer",
        "Normalised mean intensity per layer; brightness relative to per-image total."),
    "mean":                 ("Intensity (mean)", "layer",
        "Mean pixel intensity per layer, averaged across layers; higher = denser/brighter."),
    "medNormalized":        ("Intensity (median)", "layer",
        "Normalised median intensity per layer."),
    "med":                  ("Intensity (median)", "layer",
        "Median pixel intensity per layer, aggregated across layers."),
    "maxNormalized":        ("Intensity (max)", "layer",
        "Normalised maximum intensity per layer."),
    "max":                  ("Intensity (max)", "layer",
        "Maximum pixel intensity per layer; localised bright spots (clumps, active cells)."),
    "minNormalized":        ("Intensity (min)", "layer",
        "Normalised minimum intensity per layer."),
    "minPropNormalized":    ("Morphological (void fraction)", "layer",
        "Normalised void fraction: pixels below mean intensity; empty/dark regions."),
    "minProp":              ("Morphological (void fraction)", "layer",
        "Void fraction: number of pixels below mean intensity; empty regions not occupied by biofilm."),
    "minNormalized":        ("Intensity (min)", "layer",
        "Normalised minimum intensity per layer."),
    "min":                  ("Intensity (min)", "layer",
        "Minimum pixel intensity per layer; darkest/least-dense regions, gaps, heterogeneity."),
    "stdNormalized":        ("Intensity (std)", "layer",
        "Normalised standard deviation of intensity per layer."),
    "std":                  ("Intensity (std)", "layer",
        "Standard deviation of intensity per layer; low = uniform, high = high contrast/variability."),
    # ---- Texture / morphology (image-level) ----
    "Homogenity":           ("Texture (homogeneity)", "image",
        "Homogeneity from the co-occurrence matrix (Beyenal 2004); similarity of neighbouring objects, higher = more uniform."),
    "ThicknessThreshold":   ("Morphological (thickness)", "image",
        "Vertical height at which intensity drops to zero, averaged; biofilm thickness."),
    "RoughnessThreshold":   ("Morphological (roughness)", "image",
        "Average height difference between neighbouring surface pixels; surface roughness."),
    "SpreadingHorizontal":  ("Morphological (spreading)", "image",
        "Variance of biofilm pixel coordinates in x+y (Mueller 2006); horizontal spread."),
    "SpreadingVertical":    ("Morphological (spreading)", "image",
        "Variance of biofilm pixel coordinates in z; vertical spread."),
    "SpreadingTotal":       ("Morphological (spreading)", "image",
        "Summed variance in x, y and z; total spread."),
    "GPTFractalDimNormalized": ("Texture (fractal dimension)", "layer",
        "Normalised fractal dimension: geometric complexity/irregularity; higher = more branched/complex."),
    "GPTFractalDim":        ("Texture (fractal dimension)", "layer",
        "Fractal dimension: geometric complexity/space-filling; higher = more complex/branched, lower = smoother."),
    "GPTHomogeneity":       ("Texture (homogeneity)", "image",
        "Layer-wise homogeneity aggregated per image; higher = more uniform texture."),
    "GPTContrast":          ("Texture (contrast)", "image",
        "GLCM contrast: variability in neighbouring pixel intensities."),
    "GPTCorrelation":       ("Texture (correlation)", "image",
        "GLCM correlation: predictability of grey-level pairs."),
    "GPTDissimilarity":     ("Texture (dissimilarity)", "image",
        "GLCM dissimilarity: local grey-level variation (linear); opposite of homogeneity."),
    "GPTEnergy":            ("Texture (energy)", "image",
        "GLCM energy: uniformity/orderliness of texture (0-1)."),
}

# Order matters: longer / more-specific keys must be tried before shorter ones
# (e.g. 'counts(Norminten<' before 'counts(inten<', 'meanNormalized' before 'mean').
S4_KEYS_ORDERED = sorted(S4.keys(), key=len, reverse=True)

AGG_SUFFIXES = ["mean", "median", "std", "var", "min", "max", "q10", "q25", "q50",
                "q75", "q90", "q95", "sum", "skew", "kurt"]


def parse_feature(name: str):
    """Return (base_token_for_lookup, human_base, aggregation) from a raw feature string."""
    # aggregation: the trailing _<agg>.txt
    agg = ""
    m = re.search(r"_([A-Za-z0-9]+)\.txt$", name)
    if m and m.group(1) in AGG_SUFFIXES:
        agg = m.group(1)
    # base: strip the -SUMMARY... and trailing bits to get the leading token
    base_raw = re.split(r"-SUMMARY", name)[0]
    # find the S4 category by matching the datafile stem
    category, per, desc = ("Unmapped", "", "")
    matched_key = None
    for k in S4_KEYS_ORDERED:
        # normalise the key for matching (strip trailing '(' etc. handled by 'in')
        probe = k
        if probe in name:
            category, per, desc = S4[k]
            matched_key = k
            break
    return base_raw, agg, category, per, desc, matched_key


def build_hint(category, agg, per, class_medians, classes):
    """A plain-language HINT (not a conclusion) scaffolding hypothesis-building."""
    if not class_medians:
        return "Compare per-class values to see direction of difference."
    # which class highest / lowest by median
    hi = max(class_medians, key=class_medians.get)
    lo = min(class_medians, key=class_medians.get)
    # short plain gloss of what the category measures
    gloss = {
        "Texture (fractal dimension)": "geometric complexity / space-filling",
        "Morphological (void fraction)": "empty space between biofilm",
        "Morphological (volume)": "amount of biofilm",
        "Morphological (thickness)": "biofilm thickness",
        "Morphological (roughness)": "surface roughness",
        "Morphological (spreading)": "how the biofilm spreads spatially",
        "Morphological (coverage)": "surface coverage",
        "Morphological (counts)": "abundance-related structure (not absolute abundance)",
        "Intensity (mean)": "typical brightness / density",
        "Intensity (max)": "bright hotspots",
        "Intensity (min)": "dark regions / gaps",
        "Intensity (std)": "brightness variability / contrast",
        "Intensity (contrast)": "contrast between layers",
        "Intensity (global)": "overall brightness",
        "Intensity (layer-wise global)": "brightness change between layers",
        "Intensity (median)": "typical brightness",
        "Texture (homogeneity)": "texture uniformity",
        "Texture (contrast)": "texture contrast",
        "Texture (correlation)": "texture regularity",
        "Texture (dissimilarity)": "local texture variation",
        "Texture (energy)": "texture orderliness",
    }.get(category, "this structural property")
    per_txt = f" ({per}-wise)" if per in ("layer", "image") else ""
    return (f"Measures {gloss}{per_txt}, summarised by {agg or 'value'}. "
            f"Highest in {hi}, lowest in {lo}. "
            f"Consider what biological process could raise it in {hi} / lower it in {lo}.")


def cluster_features(sub_values: pd.DataFrame, threshold: float):
    """Greedy correlation clustering: |r| > threshold links features.
    Returns dict feature_name -> cluster_id (1-based), ordered so that larger
    clusters and higher-ranked members get lower ids."""
    feats = list(sub_values.columns)
    if len(feats) < 2:
        return {f: i + 1 for i, f in enumerate(feats)}
    corr = sub_values.corr(method="pearson").abs()
    # union-find
    parent = {f: f for f in feats}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for i in range(len(feats)):
        for j in range(i + 1, len(feats)):
            if corr.iloc[i, j] > threshold:
                union(feats[i], feats[j])

    # group, then assign ids by order of first appearance (feats already in rank order)
    groups = {}
    for f in feats:
        r = find(f)
        groups.setdefault(r, []).append(f)
    # order clusters by the best (earliest) rank of any member
    ordered = sorted(groups.values(), key=lambda g: min(feats.index(x) for x in g))
    mapping = {}
    for cid, g in enumerate(ordered, start=1):
        for f in g:
            mapping[f] = cid
    return mapping


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--datafile", required=True)
    ap.add_argument("--label", required=True)
    ap.add_argument("--date", required=True)
    ap.add_argument("--n", type=int, default=100, help="top-N label features to include")
    ap.add_argument("--date-n", type=int, default=100, help="top-N date features counted as a batch-effect overlap")
    ap.add_argument("--corr-threshold", type=float, default=0.8)
    ap.add_argument("--out-prefix", default="microics_feature_interpretation")
    args = ap.parse_args()

    # ---- load rankings ----
    lab = pd.read_csv(args.label, sep="\t")
    lab.columns = ["feature", "score"]
    # the date file may carry a stray 'label' row at the top - drop non-feature rows later
    dat = pd.read_csv(args.date, sep="\t")
    dat.columns = ["feature", "score"]

    # drop any row whose feature isn't a real feature string (e.g. a 'label' header artefact)
    lab = lab[lab["feature"].str.contains("SUMMARY", na=False)].reset_index(drop=True)
    dat = dat[dat["feature"].str.contains("SUMMARY", na=False)].reset_index(drop=True)

    lab = lab.sort_values("score", ascending=False).reset_index(drop=True)
    dat = dat.sort_values("score", ascending=False).reset_index(drop=True)

    top = lab.head(args.n).copy().reset_index(drop=True)
    date_top_set = set(dat.head(args.date_n)["feature"])
    date_rank = {f: i + 1 for i, f in enumerate(dat["feature"])}

    # ---- load datafile for per-class medians/IQR + correlation ----
    df = pd.read_csv(args.datafile, sep="\t")
    label_col = "label"
    classes = sorted(df[label_col].dropna().unique().tolist())

    top_feats = [f for f in top["feature"] if f in df.columns]
    missing = [f for f in top["feature"] if f not in df.columns]
    if missing:
        print(f"WARNING: {len(missing)} top features not found in datafile (skipped for values/clustering).", file=sys.stderr)
        for m in missing[:5]:
            print("   ", m, file=sys.stderr)

    values = df[top_feats].apply(pd.to_numeric, errors="coerce")

    # ---- correlation clustering on the top features ----
    clusters = cluster_features(values, args.corr_threshold)

    # ---- per-class median & IQR ----
    grouped = df.groupby(label_col)
    med = grouped[top_feats].median(numeric_only=True)
    q1 = grouped[top_feats].quantile(0.25, numeric_only=True)
    q3 = grouped[top_feats].quantile(0.75, numeric_only=True)
    iqr = q3 - q1

    # ---- assemble rows ----
    rows = []
    for _, r in top.iterrows():
        fname = r["feature"]
        base, agg, category, per, desc, key = parse_feature(fname)
        cl = clusters.get(fname, np.nan)
        class_medians = {c: (float(med.loc[c, fname]) if fname in med.columns and c in med.index else np.nan)
                         for c in classes}
        cm_clean = {c: v for c, v in class_medians.items() if pd.notna(v)}
        hint = build_hint(category, agg, per, cm_clean, classes)
        row = {
            "rank": None,  # fill after sort
            "label_score": r["score"],
            "cluster": cl,
            "feature": base,
            "aggregation": agg,
            "category": category,
            "per": per,
            "date_overlap": "YES" if fname in date_top_set else "",
            "date_rank": date_rank.get(fname, ""),
            "description": desc,
            "hint": hint,
            "_raw": fname,
        }
        for c in classes:
            row[f"median_{c}"] = class_medians.get(c, np.nan)
        for c in classes:
            row[f"IQR_{c}"] = (float(iqr.loc[c, fname])
                              if fname in iqr.columns and c in iqr.index else np.nan)
        rows.append(row)

    out = pd.DataFrame(rows)
    # sort by cluster, then by score within cluster
    out = out.sort_values(["cluster", "label_score"], ascending=[True, False]).reset_index(drop=True)
    out["rank"] = out["label_score"].rank(ascending=False, method="min").astype(int)

    # column order
    base_cols = ["rank", "label_score", "cluster", "feature", "aggregation",
                 "category", "per", "date_overlap", "date_rank"]
    med_cols = [f"median_{c}" for c in classes]
    iqr_cols = [f"IQR_{c}" for c in classes]
    tail_cols = ["description", "hint"]
    out = out[base_cols + med_cols + iqr_cols + tail_cols + ["_raw"]]

    # ---- write plain TSV ----
    tsv_path = f"{args.out_prefix}.tsv"
    out.drop(columns=["_raw"]).to_csv(tsv_path, sep="\t", index=False)
    print("wrote", tsv_path)

    # ---- write coloured Excel ----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # soft, distinguishable palette (cycled for many clusters)
    palette = [
        "E8F0FE", "E6F4EA", "FCE8E6", "FEF7E0", "F3E8FD", "E0F7FA",
        "FFF0E6", "EDE7F6", "E8F5E9", "FFF8E1", "E1F5FE", "FCE4EC",
        "F1F8E9", "FFEBEE", "E8EAF6", "E0F2F1",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Top features"

    display_cols = base_cols + med_cols + iqr_cols + tail_cols
    headers = {
        "rank": "Rank", "label_score": "Label score", "cluster": "Cluster",
        "feature": "Feature", "aggregation": "Aggreg.", "category": "Category",
        "per": "Per", "date_overlap": "Batch?", "date_rank": "Date rank",
        "description": "Description", "hint": "Hypothesis hint",
    }
    for c in classes:
        headers[f"median_{c}"] = f"med {c}"
        headers[f"IQR_{c}"] = f"IQR {c}"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="37474F")
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header row
    for j, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=j, value=headers[col])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # data rows, whole row tinted by cluster
    for i, (_, r) in enumerate(out.iterrows(), start=2):
        cl = r["cluster"]
        colour = palette[(int(cl) - 1) % len(palette)] if pd.notna(cl) else "FFFFFF"
        fill = PatternFill("solid", fgColor=colour)
        for j, col in enumerate(display_cols, start=1):
            val = r[col]
            if isinstance(val, float) and pd.notna(val):
                # show numbers with sensible precision
                if col.startswith("median_") or col.startswith("IQR_") or col == "label_score":
                    val = round(val, 4)
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = cell_font
            cell.fill = fill
            cell.border = border
            if col in ("description", "hint"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # column widths
    widths = {"rank": 6, "label_score": 11, "cluster": 8, "feature": 26,
              "aggregation": 9, "category": 24, "per": 7, "date_overlap": 8,
              "date_rank": 9, "description": 55, "hint": 55}
    for c in classes:
        widths[f"median_{c}"] = 11
        widths[f"IQR_{c}"] = 11
    for j, col in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 12)

    ws.freeze_panes = "D2"  # freeze rank/score/cluster + header
    ws.row_dimensions[1].height = 28

    # ---- second sheet: cluster summary ----
    ws2 = wb.create_sheet("Cluster summary")
    summ_headers = ["Cluster", "N features", "Representative feature", "Category",
                    "Best label score", "Any batch overlap?", "Highest class", "Lowest class"]
    for j, h in enumerate(summ_headers, start=1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ri = 2
    for cl, grp in out.groupby("cluster"):
        rep = grp.iloc[0]
        # highest/lowest class by median of representative
        cmeds = {c: rep[f"median_{c}"] for c in classes if pd.notna(rep[f"median_{c}"])}
        hi = max(cmeds, key=cmeds.get) if cmeds else ""
        lo = min(cmeds, key=cmeds.get) if cmeds else ""
        batch = "YES" if (grp["date_overlap"] == "YES").any() else ""
        vals = [int(cl), len(grp), rep["feature"], rep["category"],
                round(float(rep["label_score"]), 4), batch, hi, lo]
        colour = palette[(int(cl) - 1) % len(palette)]
        fill = PatternFill("solid", fgColor=colour)
        for j, v in enumerate(vals, start=1):
            cell = ws2.cell(row=ri, column=j, value=v)
            cell.font = cell_font
            cell.fill = fill
            cell.border = border
        ri += 1

    sw = [9, 11, 28, 24, 13, 15, 13, 13]
    for j, w in enumerate(sw, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = "A2"

    xlsx_path = f"{args.out_prefix}.xlsx"
    wb.save(xlsx_path)
    print("wrote", xlsx_path)

    # ---- console summary ----
    n_clusters = int(out["cluster"].nunique())
    n_batch = int((out["date_overlap"] == "YES").sum())
    print(f"\nSummary: {len(out)} top features -> {n_clusters} correlation clusters "
          f"(|r|>{args.corr_threshold}); {n_batch} flagged as batch-overlapping "
          f"(in top {args.date_n} of date).")
    print(f"Classes ({len(classes)}): {', '.join(map(str, classes))}")


if __name__ == "__main__":
    main()
